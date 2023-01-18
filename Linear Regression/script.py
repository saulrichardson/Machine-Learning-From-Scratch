import pandas as pd
from csv import reader
# import openpyxl
from openpyxl import load_workbook

# file = pd.read_excel("C:\\Users\\saulrichardson\\Downloads\\StaffingAnalysisTemplate1LoD.xlsx")
folderPath = "C:\\Users\\saulrichardson\\Downloads\\"
fileName = "Staffing Analysis Template 2LoD_07272022.xlsm"
fileName2 = "Staffing Analysis Template 1LoD_07272022.xlsm"

# df = pd.read_csv(path + "StaffingAnalysisTemplate1LoD.csv", skiprows = 5, header = None)

workbook = load_workbook(folderPath + fileName)

tab1 = workbook["Tab 1"]
tab2 = workbook["Tab 2"]

# with pd.ExcelFile("C:\\Users\\saulrichardson\\Downloads\\StaffingAnalysisTemplate1LoD.xlsm") as xls:
#     df1 = pd.read_excel(xls, 'Tab 1', header = None)
#     df2 = pd.read_excel(xls, 'Tab 2', header = None)

# df1.to_csv("Test.csv", index = None, header = True)

# read = reader(df2)

# tryy = load_workbook("C:\\Users\\saulrichardson\\Downloads\\StaffingAnalysisTemplate1LoD.xlsx")

# print(df2)
# print(read)
# for i in read:
#     print(i)

class Response:
    tasks = []
    confirm = {}
    lowerEnd = {}
    upperEnd = {}
    busySeason = ""
    modelCategory = ""
    role = ""
    team = ""

    def __init__(self, mc, r, t, excelInput):
        self.modelCategory = str(mc)
        self.role = str(r)
        self.team = t
        self.excel = excelInput

    def addT(self, task):
        self.tasks = self.tasks + [task]

    def addC(self, task, ans):
        self.confirm[task] = ans

    def addBound(self, tas, lower, upper):
        self.lowerEnd[tas] = lower
        self.upperEnd[tas] = upper
#         self.lowerEnd = int(lower)
#         self.upperend = int(upper)

    def collect(self, r):
        currentRow = r
        conf = self.excel.cell(row = currentRow,column = 6).value
        tk = self.excel.cell(row = currentRow,column = 4).value
        busy = self.excel.cell(row = currentRow,column = 9).value
        comments = self.excel.cell(row = currentRow,column = 10).value

        self.addT(tk)
        self.addC(tk, comments)
        self.addBound(tk, self.excel.cell(row = currentRow,column = 7).value, self.excel.cell(row = currentRow,column = 8).value)
        self.busySeason = busy
#     def __str__(self):
#         str(self.tasks)

#2LoD
print("ss")
rowStart = 8
colStart = 1
nRows = tab1.max_row
nCols = tab1.max_column
modelCat =  tab1.cell(row = 1,column = 4).value
currentRow = 8


indvResp = ["a"]
print(indvResp)
print("s")

while currentRow <= nRows:
    if tab1.cell(row = currentRow,column = 6).value != None:
        print(currentRow)
        role = tab1.cell(row = currentRow,column = 3).value
        team = tab1.cell(row = currentRow,column = 2).value
        running = Response(modelCat, role, team, tab1)
        running.collect(currentRow)

        nextRow = currentRow + 1

        while tab1.cell(row = nextRow,column = 3).value == None:
            running.collect(nextRow)
            nextRow = currentRow + 1

        indvResp.append(running)
        currentRow = nextRow

    else:
        currentRow = currentRow + 1

print(indvResp)
